import datetime
import json
import re

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

def write2excel(file_path, content: list, title=None):
    wb = Workbook()
    ws = wb.active
    if title is not None:
        ws.append(title)

    for row in content:
        ws.append(row)
    try:
        wb.save(file_path)
    except PermissionError:
        input("请确认文件是否关闭，任意键重试")
        wb.save(file_path)


def read_excel(file_path, sheet_num=0, begin_row=0):
    result_value = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    ws = wb[sheets[sheet_num]]
    for row in ws.rows:
        result_value.append([cell.value for cell in row])
    wb.close()
    return result_value[begin_row:]


def str2json(str_cont):
    return [json.loads("{{{}}}".format(i)) for i in re.findall(r"{(.*?)}", str_cont)]

def generate_code(start, end):
    today = datetime.date.today()
    today = today.strftime('%y%m%d')
    code_list = []
    for i in range(start, end+1):
        no_code = prefix + wuliaobianma + '_1500_'+ today + str(i).zfill(3)
        print(no_code)

if __name__ == '__main__':
    prefix = '1011270'
    shuliang = '72000'
    wuliaobianma = '4JB9ES53022'
    mingchengguige = '模块二极管TPA4050S-2/150mil trench/含铅锡块0.30mm'
    generate_code(1,48)